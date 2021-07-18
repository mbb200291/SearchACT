# -*- coding: utf-8 -*-
"""
Created on Sat Jul 17 17:29:21 2021

@author: mbb20
"""

from openpyxl import load_workbook

class ContactParser:
    def __init__(self, path_file):
        self.path_file = path_file
        self.sheet = load_workbook(path_file).worksheets[0]
        (self.idx_phone, self.idx_office, self.idx_department, 
         self.idx_chinese_name, self.idx_english_name, self.idx_email, 
         self.idx_cellphone) = self._get_colname_index()
        
    def _get_colname_index(self):
        # =============================================================================
        #     lookup the correct column index by impossible-to-change content
        # =============================================================================
        for row in self.sheet.iter_rows():
            list_cell = list(map(lambda x: x.value, row))
            if 1100 in list_cell:
                idx_phone = list_cell.index(1100)
            if "分子檢驗處" in list_cell:
                idx_office = list_cell.index('分子檢驗處')
            if "次世代定序部" in list_cell:
                idx_department = list_cell.index('次世代定序部') 
            if "陳華鍵" in list_cell:
                idx_chinese_name = list_cell.index('陳華鍵') 
            if "Hua-Chien Chen" in list_cell:
                idx_english_name = list_cell.index('Hua-Chien Chen')   
            if "hcchen@actgenomics.com" in list_cell:
                idx_email = list_cell.index('hcchen@actgenomics.com')     
            if '0955-622-114' in list_cell:
                idx_cellphone = list_cell.index('0955-622-114') 
        return idx_phone, idx_office,  idx_department, idx_chinese_name, idx_english_name,idx_email, idx_cellphone 
    
    def iter_contact_data(self):
        '''
        generator of excel data
        '''
        start = finish = False
        for row in sheet.iter_rows():
            list_row = list(row)
            #print(list_row[0])
            (phone, office,  department, chinese_name, english_name, email, 
            cellphone) = (list_row[x].value for x in (self.idx_phone, self.idx_office, 
                                                     self.idx_department, 
                                                     self.idx_chinese_name, 
                                                     self.idx_english_name, 
                                                     self.idx_email, self.idx_cellphone))
            #print(phone)
            if phone in [1100, 1500, 1200]:
                start = True
            if start and email == None:
                finish = True
            #print(start, finish)
            if start and not finish:
                yield (phone, office,  department, chinese_name, english_name, email, cellphone)
            elif finish:
                break
#%%
# parser = ContactParser('examples/20210715行動基因通訊錄含部門.xlsx')
# for i in parser.iter_contact_data():
#     print(i)
#%%


#%%
wb = load_workbook('examples/20210715行動基因通訊錄含部門.xlsx')
#%%
wb
#%%
sheet = wb['通訊錄']
#%%
# 以 for 迴圈逐一處理每個儲存格
def get_colname_index(sheet):
# =============================================================================
#     lookup the correct column index by impossible-to-change content
# =============================================================================
    for row in sheet.iter_rows():
        list_cell = list(map(lambda x: x.value, row))
        print(list_cell)
        if 1100 in list_cell:
            idx_phone = list_cell.index(1100)
        if "分子檢驗處" in list_cell:
            idx_office = list_cell.index('分子檢驗處')
        if "次世代定序部" in list_cell:
            idx_department = list_cell.index('次世代定序部') 
        if "陳華鍵" in list_cell:
            idx_chinese_name = list_cell.index('陳華鍵') 
        if "Hua-Chien Chen" in list_cell:
            idx_english_name = list_cell.index('Hua-Chien Chen')   
        if "hcchen@actgenomics.com" in list_cell:
            idx_email = list_cell.index('hcchen@actgenomics.com')     
        if '0955-622-114' in list_cell:
            idx_cellphone = list_cell.index('0955-622-114') 
    return idx_phone, idx_office,  idx_department, idx_chinese_name, idx_english_name,idx_email, idx_cellphone
#%%
idx_phone, idx_office,  idx_department, idx_chinese_name, idx_english_name,idx_email, idx_cellphone = get_colname_index(wb.worksheets[0])
print(idx_phone, idx_office,  idx_department, idx_chinese_name, idx_english_name,idx_email, idx_cellphone)
#%%
def excel_row_parser(sheet):
    start = finish = False
    for row in sheet.iter_rows():
        list_row = list(row)
        #print(list_row[0])
        phone, office,  department, chinese_name, english_name, email, cellphone = (
            list_row[x].value for x in (idx_phone, idx_office,  idx_department, idx_chinese_name, 
                                  idx_english_name,idx_email, idx_cellphone))
        #print(phone)
        if phone in [1100, 1500, 1200]:
            start = True
        if start and email == None:
            finish = True
        #print(start, finish)
        if start and not finish:
            yield (phone, office,  department, chinese_name, english_name, email, cellphone)
        elif finish:
            break
#%%
for r in excel_row_parser(wb.worksheets[0]):
    print(r)
#%%
import pickle
def read_pickle(path_):
    with open(path_, 'rb') as file:
        return pickle.load(file)
#%%
dic = read_pickle('D:/ben/SearchACT/_dict_data.pkl')
#%%
dic 
#%%
def min_edit_distance(str_target, str_input, cost_match=0, cost_ins=1, cost_del=1, cost_sub=2, cost_switch=2):
    matrix = [[0 for i in range(len(str_target)+1)] for j in range(len(str_input)+1)]
    for i in range(1, len(str_target)+1):
        for j in range(1, len(str_input)+1):
            min_cost = min(
                matrix[j-1][i-1] + cost_sub * (0 if str_target[i-1]==str_input[j-1] else 1), # substitution
                matrix[j][i-1] + cost_del, # delete on input
                matrix[j-1][i] + cost_ins, # insert on input
            )
            matrix[j][i] = min_cost
    return matrix[-1][-1]

#%%
min_edit_distance('abc', 'abcd')

#%%

def min_edit_distance_incswitch(str_target, str_input, cost_ins=1, cost_del=1, cost_sub=2, cost_switch=1):
    '''
    calculate Damerau–Levenshtein distance
    '''
    matrix = [[0 for i in range(len(str_target)+1)] for j in range(len(str_input)+1)]
    for i in range(len(str_target)+1):
        matrix[0][i] = i
    for j in range(len(str_input)+1):
        matrix[j][0] = j
        
    for i in range(1, (len(str_target)+1)):
        for j in range(1, len(str_input)+1):
            if str_target[i-1]==str_input[j-1]:
                cost_sub_ = 0
            else:
                cost_sub_ = cost_sub
            matrix[j][i] = min(
                matrix[j-1][i-1] + cost_sub_, # substitution
                matrix[j][i-1] + cost_del, # delete on input
                matrix[j-1][i] + cost_ins, # insert on input
            )
            if (i>2 and j>2) and str_target[i-1]==str_input[j-2] and str_target[i-2]==str_input[j-1]:
                matrix[j][i] = min(
                    matrix[j][i], 
                    matrix[j-2][i-2] + cost_switch
                    )
    return matrix[-1][-1]
#%%
min_edit_distance('abc', 'acb')
#%%
min_edit_distance('abc', ' ')
#%%
min_edit_distance_incswitch('ben', 'chen')
