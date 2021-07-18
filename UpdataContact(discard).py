'''
To create index data for SearchAct main program use.

'''

#import pandas as pd
from sys import argv
from os import path, listdir
import pickle
from openpyxl import load_workbook

class ContactParser:
    def __init__(self, path_file):
        self.path_file = path_file
        self.sheet = load_workbook(path_file).worksheets[0]
        (self.idx_phone, self.idx_office, self.idx_department, 
         self.idx_chinese_name, self.idx_english_name, self.idx_email, 
         self.idx_cellphone) = self._get_colname_index()
        print('Load excel file:', path_file)

    def _get_colname_index(self):
        # =============================================================================
        #     lookup the correct column index by impossible-to-change content
        # =============================================================================
        for row in self.sheet.iter_rows():
            list_cell = list(map(lambda x: x.value, row))
            if 1100 in list_cell:
                idx_phone = list_cell.index(1100)
            if "分子檢驗處" in list_cell:
                idx_office = list_cell.index('分子檢驗處')
            if "次世代定序部" in list_cell:
                idx_department = list_cell.index('次世代定序部') 
            if "陳華鍵" in list_cell:
                idx_chinese_name = list_cell.index('陳華鍵') 
            if "Hua-Chien Chen" in list_cell:
                idx_english_name = list_cell.index('Hua-Chien Chen')   
            if "hcchen@actgenomics.com" in list_cell:
                idx_email = list_cell.index('hcchen@actgenomics.com')     
            if '0955-622-114' in list_cell:
                idx_cellphone = list_cell.index('0955-622-114') 
        return idx_phone, idx_office,  idx_department, idx_chinese_name, idx_english_name,idx_email, idx_cellphone 
    
    def iter_contact_data(self):
        '''
        generator of excel data
        '''
        start = finish = False
        for row in self.sheet.iter_rows():
            list_row = list(row)
            #print(list_row[0])
            (phone, office,  department, chinese_name, english_name, email, 
            cellphone) = (list_row[x].value for x in (self.idx_phone, self.idx_office, 
                                                     self.idx_department, 
                                                     self.idx_chinese_name, 
                                                     self.idx_english_name, 
                                                     self.idx_email, self.idx_cellphone))
            #print(phone)
            if phone in [1100, 1500, 1200]:
                start = True
            if start and email == None:
                finish = True
            #print(start, finish)
            if start and not finish:
                yield (phone, office,  department, chinese_name, english_name, email, cellphone)
            elif finish:
                break
#%%
# parser = ContactParser('examples/20210715行動基因通訊錄含部門.xlsx')
# for i in parser.iter_contact_data():
#     print(i)


def save_obj_to_pickle(path_, obj): #will end with pickle
    print('# saving file to:', '%s.pkl'%path_)
    with open('%s.pkl'%path_, 'wb') as f_pkl:
        pickle.dump(obj, f_pkl)

def read_pickle(path_):
    with open(path_, 'rb') as file:
        return pickle.load(file)

def locate_contact_file(cwd=None):
    if not cwd:
        cwd = path.dirname(path.abspath(__file__))
    print(cwd)
    path_contact = sorted([x for x in listdir(cwd) if ".xlsx" in x and not x.startswith('~$')], key=path.getmtime)[-1]
    #print([x for x in os.listdir(cwd) if ".xlsx" in x])
    print('# Read file:', path_contact )
    return path.join(cwd, path_contact)

def remove_blank(tab):
    #First, find NaN entries in first column
    blank_row_bool = tab.iloc[:,4].isna()
    #Next, get index of first NaN entry
    blank_row_index =  [i for i, x in enumerate(blank_row_bool) if x][0]
    #Finally, restrict dataframe to rows before the first NaN entry
    return tab.iloc[:(blank_row_index)]

'''
def read_excel(path_contact):
    if path.isfile(path_contact):
        #contacts = pd.read_excel(path_contact, nrows=nrows)
        contacts = remove_blank(pd.read_excel(path_contact))
        contacts.dropna(how="all", inplace=True)
#        contacts = contacts.loc[(~contacts['信     箱'].isna()) & (contacts['信     箱'].str.contains("@"))]
        return contacts, path.splitext(path_contact)[-1]
    else:
        print('*** Contact file not exist')
        return None
'''

def make_dict(path_contact):#, _nrows):
    ## Read excel
    # contacts, excel_file_name = read_excel(path_contact)
    # if type(contacts)!=pd.core.frame.DataFrame:
    #     return None
    contacts = ContactParser(path_contact)
    excel_file_name = path.basename(path_contact)

    ## Make dictionary
    dict_to_phone = {} #
    dict_phone_Info = {'*version':excel_file_name, '*ver':excel_file_name}
    
    DI_Apartment_Abbrev = {
        "執行長室": "CEO", 
        "技術中心": "CTO", 
        "營運中心": "CFO", 
        #"Group CFO": "", 
        "研究開發處": "RD", 
        "研究開發部": "RD", 
        #"技術移轉部": "", 
        #"智慧財產部": "", 
        #"商業卓越處": "", 
        #"系統整合部": "", 
        #"產品管理部": "", 
        "生物資訊暨人工智慧處": "BI", 
        "生物資訊部": "BI", 
        "人工智慧部": "AI", 
        "數據分析部": "DA", 
        "數據智能部": "DI", 
        #"資料科學處": "", 
        #"分子檢驗處": "", 
        "次世代定序部": "NGS", 
        "轉譯醫學處": "TGI", 
        "癌症基因體部": "", 
        #"專案管理部": "PM", 
        "醫藥資訊部": "MI", 
        #"臨床醫學部": "", 
        "品保與環境監控部": "QA", 
        "法規事務處": "RA", 
        #"銷售業務處": "", 
        #"銷售業務部": "", 
        #"業務行政部": "", 
        #"臨床衛教部": "", 
        #"行政資源處": "", 
        #"行政資源部": "", 
        #"會計處": "", 
        #"財務與出納部": "", 
        "事業暨企業發展處": "BD", 
        "資訊處": "IT", 
        "人力資源處": "HR", 
    }
    #for i,R in contacts.iterrows():
    for (phone, office,  department, chinese_name, english_name, email, 
    cellphone) in contacts.iter_contact_data():
        english_name_var2 = english_name.replace('-', ' ').strip()
        english_name_var3 = english_name.replace('-', '').strip()
        
        print(chinese_name, english_name)
        
        key_ = email.split('@')[0]
        ## first part: key_ to result
        dict_phone_Info[key_] = [str(x) for x in [chinese_name, english_name, department, email, phone, cellphone]]

        ## second part: query items to key
        dict_to_phone.setdefault(chinese_name, set()).add(key_)
        #dict_to_phone.setdefault(chinese_name[:-1], set()).add(key_)
        #dict_to_phone.setdefault(chinese_name[-1], set()).add(key_)
        dict_to_phone.setdefault(chinese_name[0], set()).add(key_)
        dict_to_phone.setdefault(english_name, set()).add(key_)
        dict_to_phone.setdefault(english_name.lower(), set()).add(key_) # lower
        dict_to_phone.setdefault(english_name_var2, set()).add(key_)
        dict_to_phone.setdefault(english_name_var2.lower(), set()).add(key_)  # lower
        dict_to_phone.setdefault(english_name_var3, set()).add(key_)
        dict_to_phone.setdefault(english_name_var3.lower(), set()).add(key_)  # lower
        dict_to_phone.setdefault(''.join([x[0].upper() for x in english_name.split()]), set()).add(key_)     # abbrev eng name (eg, Shu-Jen Chen => SJC)
        dict_to_phone.setdefault(''.join([x[0].upper() for x in english_name_var2.split()]), set()).add(key_) # abbrev eng name
        dict_to_phone.setdefault(''.join([x[0].upper() for x in english_name_var3.split()]), set()).add(key_) # abbrev eng name

        dict_to_phone.setdefault(english_name.split()[-1][0].upper()+''.join([x[0].upper() for x in english_name.split()[:-1]]), set()).add(key_)      # abbrev eng name (eg, Shu Jen Chen => CSJ)
        dict_to_phone.setdefault(english_name.split()[-1][0].upper()+''.join([x[0].upper() for x in english_name_var2.split()[:-1]]), set()).add(key_) # abbrev eng name
        dict_to_phone.setdefault(english_name.split()[-1][0].upper()+''.join([x[0].upper() for x in english_name_var3.split()[:-1]]), set()).add(key_) # abbrev eng name 

        # email user id
        dict_to_phone.setdefault(email.split('@')[0].lower(), set()).add(key_)
        # by Phone num
        dict_to_phone.setdefault(str(phone), set()).add(key_)
        # by cell phone num
        dict_to_phone.setdefault(str(cellphone), set()).add(key_)
        # by department
        dict_to_phone.setdefault(str(department), set()).add(key_)
        ## by department's abbreviation
        dict_to_phone.setdefault(DI_Apartment_Abbrev.get(str(department), str(department)), set()).add(key_)


    #print(os.path.split(sys.path[0])[0])
    path_term_key = path.join(path.dirname(path_contact), '_dict_terms_key')
    path_key_contact = path.join(path.dirname(path_contact), '_dict_key_contacts')
    # path_data = path.join(path.dirname(path_contact), '_dict_data')

    #save_obj_to_pickle(path_data, [dict_to_phone, dict_phone_Info])
    save_obj_to_pickle(path_term_key, dict_to_phone)
    save_obj_to_pickle(path_key_contact, dict_phone_Info)


def main():
    print('# Loading')
    make_dict(locate_contact_file())# int(sys.argv[2]))
    print('# ================================== Finish ================================== ')


if __name__ == '__main__':
    main()
