"""
To manage dictionary update operations.

"""
import os
import pathlib
import pickle
import sys
import unicodedata
from os import listdir
from os import path
from typing import Dict
from typing import Union

from openpyxl import load_workbook

from .crud import get_department_all
from .install import main as init_db


class Contact:
    def __init__(self, cwd=None):
        self._cwd = cwd or (
            os.path.dirname(sys.executable)
            if getattr(sys, "frozen", False)
            else pathlib.Path(__file__).parent.parent.resolve()
        )
        self._path_file = self.locate_latest_contact_file(cwd=self.cwd)
        self._col_index = {
            "idx_phone": 0,
            "idx_office": 1,
            "idx_department": 2,
            "idx_chinese_name": 3,
            "idx_english_name": 4,
            "idx_email": 5,
            "idx_cellphone": 6,
        }
        self._data = self.read_pickle(path.join(self.cwd, "_dict_data.pkl"))

    @property
    def cwd(self):
        return self._cwd

    @property
    def path_file(self):
        return self._path_file

    @path_file.setter
    def path_file(self, path_file: str):
        self._path_file = path_file

    @property
    def col_index(self):
        return self._col_index

    @col_index.setter
    def col_index(self, mapping_of_contact_cols: dict):
        self._col_index = mapping_of_contact_cols

    @property
    def data(self):
        return self._data

    @data.setter
    def data(self, dict_mapping_data: dict):
        self._data = dict_mapping_data

    def _parse_contact(self) -> Dict:

        return dict(
            zip(
                (
                    "idx_phone",
                    "idx_office",
                    "idx_department",
                    "idx_chinese_name",
                    "idx_english_name",
                    "idx_email",
                    # "idx_cellphone"
                ),
                self._get_colname_index(),
            )
        )

    def _get_colname_index(self):
        # =============================================================================
        #     lookup the correct column index by impossible-to-change content
        # =============================================================================
        sheet = load_workbook(self.path_file).worksheets[0]
        for row in sheet.iter_rows():
            list_cell = list(map(lambda x: x.value, row))
            if 1100 in list_cell:
                idx_phone = list_cell.index(1100)
            if "分子檢驗處" in list_cell:
                idx_office = list_cell.index("分子檢驗處")
            if "次世代定序部" in list_cell:
                idx_department = list_cell.index("次世代定序部")
            if "陳華鍵" in list_cell:
                idx_chinese_name = list_cell.index("陳華鍵")
            if "Hua-Chien Chen" in list_cell:
                idx_english_name = list_cell.index("Hua-Chien Chen")
            if "hcchen@actgenomics.com" in list_cell:
                idx_email = list_cell.index("hcchen@actgenomics.com")
            # if "0955-622-114" in list_cell:
            #     idx_cellphone = list_cell.index("0955-622-114")
        return (
            idx_phone,
            idx_office,
            idx_department,
            idx_chinese_name,
            idx_english_name,
            idx_email,
            # idx_cellphone,
        )

    def locate_latest_contact_file(self, cwd: str) -> Union[None, str]:

        name_contact = None

        li_name_contact = sorted(
            [x for x in listdir(cwd) if ".xlsx" in x and not x.startswith("~$")],
            key=path.getmtime,
        )

        if li_name_contact:
            name_contact = li_name_contact[-1]
        # else:
        #     print("[ERROR] xlsx files not found!, using default version of contact.")

        return path.join(cwd, name_contact) if name_contact else name_contact

    def iter_contact_data(self):
        """
        generator of excel data
        """
        sheet = load_workbook(self.path_file).worksheets[0]
        start = finish = False
        for row in sheet.iter_rows():
            list_row = list(row)
            # print(list_row)
            (
                phone,
                office,
                department,
                chinese_name,
                english_name,
                email,
                # cellphone,
            ) = (list_row[x].value for x in (self._col_index.values()))
            # print(phone)
            if phone in [1100, 1500, 1200]:
                start = True
            if start and email is None:
                finish = True
            # print(start, finish)
            if start and not finish:
                yield (
                    phone,
                    unicodedata.normalize("NFKC", office),
                    unicodedata.normalize("NFKC", department),
                    unicodedata.normalize("NFKC", chinese_name),
                    unicodedata.normalize("NFKC", english_name),
                    unicodedata.normalize("NFKC", email),
                    # cellphone,
                )
            elif finish:
                break

    def _key_string_modify(self, key_string):
        return "@%s@" % "@".join(list(key_string.strip()))

    def rebuild_dict(self, cwd):
        self.path_file = self.locate_latest_contact_file(cwd=cwd)
        self.make_dict()

    def make_dict(self) -> dict:
        assert (
            self.path_file is not None
        ), "檔案不存在，請先設定檔案路徑。 e.g self.path_file = 'some_path_to_file'"

        self.col_index = self._parse_contact()

        excel_file_name = path.basename(self.path_file)

        init_db()

        # Make dictionary
        dict_mapping_data = {"*version": excel_file_name, "*ver": excel_file_name}

        for (
            phone,
            office,
            department,
            chinese_name,
            english_name,
            email,
            # cellphone,
        ) in self.iter_contact_data():
            english_name_var2 = english_name.replace("-", " ").strip()
            english_name_var3 = english_name.replace("-", "").strip()

            key_ = self._key_string_modify(email.split("@")[0])

            # first part: key_ to result
            dict_mapping_data[key_] = [
                str(x) for x in [chinese_name, english_name, department, email, phone]
            ]

            # second part: query items to key
            dict_mapping_data.setdefault(chinese_name, set()).add(key_)
            dict_mapping_data.setdefault(chinese_name[0], set()).add(key_)
            dict_mapping_data.setdefault(english_name, set()).add(key_)
            dict_mapping_data.setdefault(english_name.lower(), set()).add(key_)  # lower
            dict_mapping_data.setdefault(english_name_var2, set()).add(key_)
            dict_mapping_data.setdefault(english_name_var2.lower(), set()).add(
                key_
            )  # lower
            dict_mapping_data.setdefault(english_name_var3, set()).add(key_)
            dict_mapping_data.setdefault(english_name_var3.lower(), set()).add(
                key_
            )  # lower
            dict_mapping_data.setdefault(
                "".join([x[0].upper() for x in english_name.split()]), set()
            ).add(
                key_
            )  # abbrev eng name (eg, Shu-Jen Chen => SJC)
            dict_mapping_data.setdefault(
                "".join([x[0].upper() for x in english_name_var2.split()]), set()
            ).add(
                key_
            )  # abbrev eng name
            dict_mapping_data.setdefault(
                "".join([x[0].upper() for x in english_name_var3.split()]), set()
            ).add(
                key_
            )  # abbrev eng name

            dict_mapping_data.setdefault(
                english_name.split()[-1][0].upper()
                + "".join([x[0].upper() for x in english_name.split()[:-1]]),
                set(),
            ).add(
                key_
            )  # abbrev eng name (eg, Shu Jen Chen => CSJ)
            dict_mapping_data.setdefault(
                english_name.split()[-1][0].upper()
                + "".join([x[0].upper() for x in english_name_var2.split()[:-1]]),
                set(),
            ).add(
                key_
            )  # abbrev eng name
            dict_mapping_data.setdefault(
                english_name.split()[-1][0].upper()
                + "".join([x[0].upper() for x in english_name_var3.split()[:-1]]),
                set(),
            ).add(
                key_
            )  # abbrev eng name

            # email user id
            dict_mapping_data.setdefault(email.split("@")[0].lower(), set()).add(key_)
            # by Phone num
            dict_mapping_data.setdefault(str(phone), set()).add(key_)
            # by cell phone num
            # dict_mapping_data.setdefault(str(cellphone), set()).add(key_)
            # dict_mapping_data.setdefault(
            #     str(cellphone).replace("-", "").replace(" ", ""), set()
            # ).add(key_)
            # by department
            dict_mapping_data.setdefault(str(department), set()).add(key_)
            # by department's abbreviation
            dict_mapping_data.setdefault(
                {i.name: i.abbreviation for i in get_department_all()}.get(
                    str(department), str(department)
                ),
                set(),
            ).add(key_)
            # tokenized name
            for n in (english_name, english_name_var2, english_name_var3, chinese_name):
                for t in n.split(" "):
                    dict_mapping_data.setdefault(t.lower(), set()).add(key_)
                    dict_mapping_data.setdefault(t, set()).add(key_)

        path_data = path.join(self.cwd, "_dict_data")
        self.save_pickle(path_data, dict_mapping_data)
        return dict_mapping_data

    def save_pickle(self, path_, obj):  # will end with pickle
        print("# saving file to:", "%s.pkl" % path_)
        with open("%s.pkl" % path_, "wb") as f_pkl:
            pickle.dump(obj, f_pkl)

    def read_pickle(self, path_):
        assert os.path.isfile(path_), "pickle檔案不存在，請放在同一資料夾下。"
        with open(path_, "rb") as file:
            return pickle.load(file)

    def check_version(self):
        try:
            self._path_file = self.locate_latest_contact_file(cwd=self.cwd)

            if not self.path_file:
                return "檔案不存在，請先將xlsx放在同一目錄下。"

            excel_version = path.basename(self.path_file)

            if excel_version != "" and self.data.get("*version") != excel_version:
                print("new version excel found, re-build _dict_data.pkl.")
                _data = self.make_dict()
                self.data = _data
                return "new version excel found, update finished."

            return "no need to update."

        except Exception as e:
            return e

    def key_exist(self, str_input_KEY):
        return self.key_string_modify(str_input_KEY)

    def add_searchTerm_to_key(self, str_input_Term, str_input_KEY):
        self.data.setdefault(str_input_Term, set()).add(
            self._key_string_modify(str_input_KEY)
        )
        self.save_pickle(path.splitext(self.path_file)[0], self.data)
        print(f'Successfuly add new search terms. "{str_input_Term}":"{str_input_KEY}"')
        return 0

    def add_contact_info(self, str_input_KEY, str_add_info):
        self.data[self._key_string_modify(str_input_KEY)].append(str_add_info)
        self.save_pickle(path.splitext(self.path_file)[0], self.data)
        print(
            f'Successfuly add new info terms. "{self.data[self._key_string_modify(str_input_KEY)]}"'
        )
        return 0
